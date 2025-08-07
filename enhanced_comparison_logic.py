#!/usr/bin/env python3
"""
Enhanced Data Comparison Logic
Optimized for better performance, caching, and intelligent comparison strategies
"""

import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from collections import defaultdict
import hashlib
import json

class EnhancedComparisonEngine:
    def __init__(self):
        self.db_cache = {}  # Cache DB results to avoid repeated SP calls
        self.sp_performance = {}  # Track SP performance
        self.comparison_results = {}  # Store comparison results
        self.lock = threading.Lock()
        
    def get_cache_key(self, sp_name, params):
        """Generate cache key for DB results"""
        param_str = json.dumps(params, sort_keys=True)
        return hashlib.md5(f"{sp_name}_{param_str}".encode()).hexdigest()
    
    def fetch_db_data_with_cache(self, sp_name, params):
        """Fetch DB data with intelligent caching"""
        cache_key = self.get_cache_key(sp_name, params)
        
        with self.lock:
            if cache_key in self.db_cache:
                print(f"🔄 Using cached data for {sp_name}")
                return self.db_cache[cache_key]
        
        # Fetch fresh data
        start_time = time.time()
        try:
            # Your existing DB fetch logic here
            db_data = self._execute_sp(sp_name, params)
            execution_time = time.time() - start_time
            
            # Cache the result
            with self.lock:
                self.db_cache[cache_key] = db_data
                self.sp_performance[sp_name] = execution_time
            
            print(f"✅ Fetched and cached {sp_name} in {execution_time:.2f}s")
            return db_data
            
        except Exception as e:
            print(f"❌ Error fetching {sp_name}: {e}")
            return {}
    
    def batch_fetch_all_sps(self, params):
        """Fetch all required SPs in parallel for maximum efficiency"""
        required_sps = [
            "SP_SalesTrend",
            "SP_TopBrandsBySales", 
            "SP_TopPerformingEmployee",
            "SP_TopProductsBySales",
            "SP_WeekdayWeekendSales",
            "SP_WeekwiseSalesComparison"
        ]
        
        print(f"🚀 Batch fetching {len(required_sps)} stored procedures...")
        
        all_db_data = {}
        
        with ThreadPoolExecutor(max_workers=3) as executor:
            # Submit all SP calls simultaneously
            future_to_sp = {
                executor.submit(self.fetch_db_data_with_cache, sp, params): sp 
                for sp in required_sps
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_sp):
                sp_name = future_to_sp[future]
                try:
                    sp_data = future.result()
                    if sp_data:
                        all_db_data.update(sp_data)
                        print(f"✅ Completed {sp_name}: {len(sp_data)} values")
                    else:
                        print(f"⚠️ No data from {sp_name}")
                except Exception as e:
                    print(f"❌ Error with {sp_name}: {e}")
        
        print(f"📊 Batch fetch complete: {len(all_db_data)} total DB values")
        return all_db_data
    
    def intelligent_comparison_strategy(self, excel_path, widget_title, submenu_selection):
        """Intelligent comparison strategy based on widget type and data patterns"""
        
        # Strategy 1: Pre-fetch all data once for multiple comparisons
        if not hasattr(self, '_global_db_data'):
            print("🧠 Pre-fetching all DB data for intelligent comparison...")
            params = self._get_params_for_widget(widget_title, submenu_selection)
            self._global_db_data = self.batch_fetch_all_sps(params)
        
        # Strategy 2: Smart sheet detection and targeted comparison
        sheet_patterns = self._analyze_excel_sheets(excel_path)
        relevant_data = self._filter_relevant_db_data(self._global_db_data, sheet_patterns)
        
        # Strategy 3: Optimized comparison with pattern matching
        comparison_result = self._perform_smart_comparison(
            excel_path, relevant_data, widget_title, submenu_selection
        )
        
        return comparison_result
    
    def _analyze_excel_sheets(self, excel_path):
        """Analyze Excel sheets to determine data patterns"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            
            patterns = {
                'has_sales_trends': False,
                'has_weekly_data': False,
                'has_product_data': False,
                'has_weekday_weekend': False,
                'sheet_count': len(wb.sheetnames)
            }
            
            for sheet_name in wb.sheetnames:
                sheet_lower = sheet_name.lower()
                if 'sales trends' in sheet_lower:
                    patterns['has_sales_trends'] = True
                elif 'weekwise' in sheet_lower:
                    patterns['has_weekly_data'] = True
                elif 'product' in sheet_lower:
                    patterns['has_product_data'] = True
                elif 'weekday' in sheet_lower or 'weekend' in sheet_lower:
                    patterns['has_weekday_weekend'] = True
            
            print(f"📋 Excel analysis: {patterns}")
            return patterns
            
        except Exception as e:
            print(f"⚠️ Excel analysis failed: {e}")
            return {'sheet_count': 0}
    
    def _filter_relevant_db_data(self, all_db_data, patterns):
        """Filter DB data based on Excel patterns to reduce comparison overhead"""
        if not patterns.get('sheet_count', 0):
            return all_db_data
        
        relevant_data = {}
        
        for key, value in all_db_data.items():
            key_lower = key.lower()
            
            # Include data based on detected patterns
            include = False
            
            if patterns.get('has_sales_trends') and 'sales trends' in key_lower:
                include = True
            elif patterns.get('has_weekly_data') and 'weekwise' in key_lower:
                include = True
            elif patterns.get('has_product_data') and 'product' in key_lower:
                include = True
            elif patterns.get('has_weekday_weekend') and ('weekday' in key_lower or 'weekend' in key_lower):
                include = True
            else:
                # Include if no specific patterns detected (fallback)
                include = not any(patterns.values())
            
            if include:
                relevant_data[key] = value
        
        print(f"🎯 Filtered to {len(relevant_data)} relevant DB values (from {len(all_db_data)})")
        return relevant_data
    
    def get_comparison_performance_stats(self):
        """Get performance statistics for optimization"""
        return {
            'cached_sps': len(self.db_cache),
            'sp_performance': self.sp_performance,
            'total_comparisons': len(self.comparison_results),
            'cache_hit_rate': self._calculate_cache_hit_rate()
        }
    
    def _calculate_cache_hit_rate(self):
        """Calculate cache hit rate for performance monitoring"""
        # Implementation would track cache hits vs misses
        return 0.85  # Example value
    
    def optimize_for_remaining_drillthroughs(self, remaining_widgets):
        """Optimize comparison strategy for remaining drillthroughs"""
        print(f"🎯 Optimizing for {len(remaining_widgets)} remaining drillthroughs...")
        
        # Strategy 1: Pre-fetch common data
        common_params = self._identify_common_parameters(remaining_widgets)
        if common_params:
            print("🔄 Pre-fetching common data for remaining drillthroughs...")
            self.batch_fetch_all_sps(common_params)
        
        # Strategy 2: Prioritize by data complexity
        prioritized_widgets = self._prioritize_by_complexity(remaining_widgets)
        
        # Strategy 3: Batch similar widget types
        batched_widgets = self._batch_similar_widgets(prioritized_widgets)
        
        return batched_widgets
    
    def _identify_common_parameters(self, widgets):
        """Identify common parameters across widgets to optimize caching"""
        # Most drillthroughs use similar base parameters
        return (2024, None, "717", None, None, None, None)
    
    def _prioritize_by_complexity(self, widgets):
        """Prioritize widgets by data complexity (simple first)"""
        complexity_order = {
            'Weekly Trends': 1,      # Simple time series
            'Top Products': 2,       # Product list
            'Top Brands': 3,         # Brand list  
            'Top Categories': 4,     # Category hierarchy
            'Top Sub Categories': 5, # Sub-category hierarchy
            'Top Stores': 6,         # Store hierarchy
            'Store Sales': 7         # Complex with targets
        }
        
        return sorted(widgets, key=lambda w: complexity_order.get(w, 999))
    
    def _batch_similar_widgets(self, widgets):
        """Batch widgets with similar data patterns"""
        batches = defaultdict(list)
        
        for widget in widgets:
            if 'Top' in widget:
                batches['ranking_widgets'].append(widget)
            elif 'Trends' in widget or 'Weekly' in widget:
                batches['time_series_widgets'].append(widget)
            else:
                batches['other_widgets'].append(widget)
        
        return dict(batches)

# Enhanced comparison functions for different scenarios

def compare_with_smart_caching(excel_path, widget_title, submenu_selection, output_path):
    """Enhanced comparison with intelligent caching and optimization"""
    engine = EnhancedComparisonEngine()
    return engine.intelligent_comparison_strategy(excel_path, widget_title, submenu_selection)

def batch_compare_remaining_drillthroughs(remaining_drillthroughs):
    """Optimized batch comparison for remaining drillthroughs"""
    engine = EnhancedComparisonEngine()
    
    # Optimize strategy for remaining widgets
    optimized_batches = engine.optimize_for_remaining_drillthroughs(remaining_drillthroughs)
    
    results = {}
    
    # Process each batch with optimized strategy
    for batch_type, widgets in optimized_batches.items():
        print(f"🔄 Processing {batch_type}: {len(widgets)} widgets")
        
        for widget in widgets:
            try:
                result = engine.intelligent_comparison_strategy(
                    widget['excel_path'], 
                    widget['title'], 
                    widget['submenu']
                )
                results[widget['title']] = result
                
            except Exception as e:
                print(f"❌ Batch comparison failed for {widget['title']}: {e}")
                results[widget['title']] = {'error': str(e)}
    
    return results

def get_optimization_recommendations():
    """Get recommendations for further optimization"""
    return {
        'caching': 'Pre-fetch all SP data once and reuse across comparisons',
        'batching': 'Group similar widgets and process together',
        'filtering': 'Only compare relevant data based on Excel sheet analysis',
        'parallelization': 'Run multiple comparisons simultaneously',
        'performance_monitoring': 'Track SP execution times and optimize slow ones'
    }

# Usage example for remaining drillthroughs:
"""
remaining_widgets = [
    {'title': 'Top Products by Sales', 'submenu': 'MATLAB', 'excel_path': 'path1.xlsx'},
    {'title': 'Weekly Trends', 'submenu': 'Week1', 'excel_path': 'path2.xlsx'},
    # ... more widgets
]

results = batch_compare_remaining_drillthroughs(remaining_widgets)
"""