from openpyxl import Workbook,load_workbook
import os
from dataBase import DatabaseConnector
from collections import defaultdict
from openpyxl import Workbook    
from dynamic_comparison_engine import DynamicComparisonEngine
from config_loader import config_loader

# Import dynamic comparison engine for enhanced functionality
try:
    DYNAMIC_ENGINE_AVAILABLE = True
    print("🧠 Dynamic comparison engine available for widget comparisons")
except ImportError:
    DYNAMIC_ENGINE_AVAILABLE = False
    print("⚠️ Dynamic comparison engine not available, using legacy methods")

databaseconnect = DatabaseConnector()
conn = databaseconnect.connect()
cursor = conn.cursor()

# Set your filter parameters
year = 2024
month = None
store = None
state = None
channel = None
fromdate = None
todate = None
params = (year, month, store, state, channel, fromdate, todate)

# Map: sheet name → stored procedure
widget_sp_map = config_loader.get_landing_page_widgets()


def normalize(name: str) -> str:
    return ''.join(name.lower().strip().replace('_', '').replace(' ', ''))


def read_widget_values(excel_path, normalized_map):
    wb = load_workbook(excel_path, data_only=True)
    widget_values = {}

    for sheet_name in wb.sheetnames:
        normalized_sheet = normalize(sheet_name)
        widget_name = normalized_map.get(normalized_sheet)

        # Fallback: try partial match
        if not widget_name:
            for norm_name, key in normalized_map.items():
                if norm_name in normalized_sheet:
                    widget_name = key
                    break

        if not widget_name:
            print(f"⚠️ No matching widget name for sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]

        if "identifier" in headers and "actual sales" in headers:
            id_idx = headers.index("identifier")
            val_idx = headers.index("actual sales")
            for row in ws.iter_rows(min_row=2, values_only=True):
                identifier = str(row[id_idx]).strip() if row[id_idx] else ""
                value = str(row[val_idx]).strip() if row[val_idx] is not None else "0"
                widget_values[f"{sheet_name.strip()} - {identifier}"] = value

        elif "identifier" in headers and "targets" in headers and "actual sales" in headers:
            id_idx = headers.index("identifier")
            actual_idx = headers.index("actual sales")
            target_idx = headers.index("targets")
            for row in ws.iter_rows(min_row=2, values_only=True):
                identifier = str(row[id_idx]).strip() if row[id_idx] else ""
                if row[actual_idx] is not None:
                    widget_values[f"{sheet_name.strip()} - {identifier}"] = str(row[actual_idx]).strip()
                if row[target_idx] is not None:
                    widget_values[f"{sheet_name.strip()} Target - {identifier}"] = str(row[target_idx]).strip()

        elif "week" in headers and "metric" in headers and "value" in headers:
            week_idx = headers.index("week")
            metric_idx = headers.index("metric")
            value_idx = headers.index("value")
            for row in ws.iter_rows(min_row=2, values_only=True):
                week = str(row[week_idx]).strip() if row[week_idx] else "Unknown"
                metric = str(row[metric_idx]).strip() if row[metric_idx] else ""
                val = str(row[value_idx]).strip() if row[value_idx] is not None else "0"
                widget_values[f"{sheet_name.strip()} - {week} {metric}"] = val

        elif "identifier/label" in headers and "metric" in headers and "excel value" in headers:
            id_idx = headers.index("identifier/label")
            metric_idx = headers.index("metric")
            val_idx = headers.index("excel value")
            for row in ws.iter_rows(min_row=2, values_only=True):
                identifier = str(row[id_idx]).strip() if row[id_idx] else "Unknown"
                metric = str(row[metric_idx]).strip() if row[metric_idx] else ""
                value = str(row[val_idx]).strip() if row[val_idx] is not None else "0"
                key = f"{widget_name} - {identifier} {metric}"
                widget_values[key] = value

        elif "previous year" in headers and "current year" in headers:
            prev_idx = headers.index("previous year")
            curr_idx = headers.index("current year")
            for row in ws.iter_rows(min_row=2, values_only=True):
                week_label = str(row[0]).strip() if row[0] else "Unknown"
                if row[prev_idx] is not None:
                    widget_values[f"{sheet_name.strip()} - {week_label} Previous"] = str(row[prev_idx]).strip()
                if row[curr_idx] is not None:
                    widget_values[f"{sheet_name.strip()} - {week_label} Current"] = str(row[curr_idx]).strip()

        else:
            print(f"⚠️ Unknown format in sheet: {sheet_name}")

    return widget_values


def fetch_db_widget_values(widget_sp_map, params):
    db_widget_data = {}

    for display_name, sp in widget_sp_map.items():
        try:
            cursor.execute(f"EXEC [{sp}] ?, ?, ?, ?, ?, ?, ?", params)
            result = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]
            print(f"Columns for {sp}: {columns}")  # Debug: Log column names

            identifier_col = value_col = target_col = week_col = prev_col = curr_col = None

            for col in columns:
                col_lower = col.lower()
                if "identifier" in col_lower or "store" in col_lower:
                    identifier_col = col
                elif any(x in col_lower for x in ["actual", "actual sales", "sales", "actual_sales"]):
                    value_col = col
                elif "target" in col_lower:
                    target_col = col
                elif "week" in col_lower:
                    week_col = col
                elif "previous" in col_lower:
                    prev_col = col
                elif "current" in col_lower:
                    curr_col = col

            if identifier_col and value_col:
                for row in result:
                    row_dict = dict(zip(columns, row))
                    identifier = row_dict.get(identifier_col)
                    actual_value = row_dict.get(value_col)
                    print(f"Row for {sp}: {row_dict}")  # Debug: Log raw row data
                    if identifier and actual_value is not None:
                        key = f"{display_name.strip()} - {str(identifier).strip()}"
                        db_widget_data[key] = str(actual_value).strip()
                    if target_col and row_dict.get(target_col) is not None:
                        target_key = f"{display_name.strip()} - {str(identifier).strip()} Target"
                        db_widget_data[target_key] = str(row_dict[target_col]).strip()

            elif identifier_col and "metric" in [c.lower() for c in columns] and "excel value" in [c.lower() for c in columns]:
                metric_col = next(c for c in columns if c.lower() == "metric")
                value_col = next(c for c in columns if c.lower() == "excel value")

                for row in result:
                    row_dict = dict(zip(columns, row))
                    identifier = str(row_dict.get(identifier_col)).strip()
                    metric = str(row_dict.get(metric_col)).strip()
                    value = row_dict.get(value_col)
                    print(f"Row for {sp}: {row_dict}")  # Debug: Log raw row data

                    if identifier and metric and value is not None:
                        # Special handling for SP_StorewiseActualVsTarget_Vertical_SortedByActual
                        if sp == "SP_StorewiseActualVsTarget_Vertical_SortedByActual" and metric.lower() == "actual sales":
                            key = f"{display_name.strip()} - {identifier}"
                        elif sp == "SP_StorewiseActualVsTarget_Vertical_SortedByActual" and metric.lower() == "target":
                            key = f"{display_name.strip()} - {identifier} Target"
                        else:
                            key = f"{display_name.strip()} - {identifier} {metric}"
                        db_widget_data[key] = str(value).strip()

            else:
                db_widget_data[display_name.strip()] = str(result[0][0]).strip() if result and result[0] else "No Data"

        except Exception as e:
            print(f"❌ Error fetching DB value for {sp}: {e}")
            db_widget_data[display_name.strip()] = "Error"

    return db_widget_data

    
def compare_widget_data(excel_path, db_data, output_path, widget_sp_map):
    def safe_round(value):
        try:
            clean_val = str(value).replace(",", "").replace("$", "").replace("₹", "").replace("K", "").replace("M", "")
            return round(float(clean_val), 2)
        except:
            return value

    def normalize(name: str) -> str:
        return ''.join(str(name).lower().strip().replace("_", "").replace(" ", ""))

    if not os.path.exists(excel_path):
        print("❌ Excel file not found.")
        return

    normalized_widget_map = {normalize(v): v for v in widget_sp_map.values()}
    reverse_widget_map = {normalize(k): k for k in widget_sp_map.keys()}

    wb_source = load_workbook(excel_path, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in wb_source.sheetnames:
        print(f"\n📄 Comparing sheet: {sheet_name}")
        ws = wb_source[sheet_name]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        col_indices = {h: i for i, h in enumerate(headers)}

        id_idx = col_indices.get("identifier") or col_indices.get("identifier/label") or 0
        actual_idx = col_indices.get("actual sales")
        target_idx = col_indices.get("targets") or col_indices.get("target")
        prev_idx = col_indices.get("previous year")
        curr_idx = col_indices.get("current year")
        metric_idx = col_indices.get("metric")
        excelval_idx = col_indices.get("excel value")

        output = []

        normalized_sheet = normalize(sheet_name)
        widget_name = None
        for norm_key, display_key in reverse_widget_map.items():
            if norm_key in normalized_sheet:
                widget_name = display_key
                break

        if not widget_name:
            print(f"⚠️ Widget name not mapped for sheet '{sheet_name}', using sheet name as fallback.")
            widget_name = sheet_name

        for row in ws.iter_rows(min_row=2, values_only=True):
            identifier = str(row[id_idx]).strip() if row[id_idx] else "Unknown"

            def compare_metric(label, col_index, suffix=""):
                if col_index is not None and row[col_index] is not None:
                    excel_val = row[col_index]
                    key = f"{widget_name} - {identifier}" if label == "Actual Sales" else f"{widget_name} - {identifier} {suffix}"
                    db_val = db_data.get(key, "Not Found")
                    
                    # Debug: If not found, try alternative key formats for drillthrough data
                    if db_val == "Not Found":
                        alternative_keys = [
                            f"Weekly Trends - {identifier} {suffix}",
                            f"Sales Summary_Weekday Weekend - {identifier} {suffix}",
                            f"Weekly Trends - {identifier.replace(' SALES', '')} {suffix}",
                            f"Weekly Trends - {identifier.replace(' SALES', '')} SALES {suffix}"
                        ]
                        
                        for alt_key in alternative_keys:
                            if alt_key in db_data:
                                print(f"✅ Found alternative key: {alt_key} (was looking for: {key})")
                                db_val = db_data[alt_key]
                                break
                        
                        if db_val == "Not Found":
                            print(f"❌ DB value not found for key: {key}")
                            print(f"🔍 Available keys containing '{identifier}': {[k for k in db_data.keys() if identifier in k][:3]}")
                    
                    excel_rounded = safe_round(excel_val)
                    db_rounded = safe_round(db_val)
                    status = "Match" if excel_rounded == db_rounded else "Mismatch"
                    output.append([identifier, label, excel_rounded, db_rounded, status])

            compare_metric("Actual Sales", actual_idx)
            compare_metric("Target", target_idx, "Target")
            compare_metric("Previous Year", prev_idx, "Previous Year")
            compare_metric("Current Year", curr_idx, "Current Year")

            if metric_idx is not None and excelval_idx is not None:
                metric = str(row[metric_idx]).strip() if row[metric_idx] else ""
                excel_val = row[excelval_idx]
                key = f"{widget_name} - {identifier} {metric}"
                db_val = db_data.get(key, "Not Found")
                excel_rounded = safe_round(excel_val)
                db_rounded = safe_round(db_val)
                status = "Match" if excel_rounded == db_rounded else "Mismatch"
                if db_val == "Not Found":
                    print(f"❌ DB value not found for key: {key}")
                output.append([identifier, metric, excel_rounded, db_rounded, status])

        if output:
            ws_out = wb_out.create_sheet(title=sheet_name[:31])
            ws_out.append(["Identifier/Label", "Metric", "Excel Value", "DB Value", "Status"])
            for line in output:
                ws_out.append(line)
            print(f"✅ Sheet '{sheet_name}' compared with {len(output)} rows.")
        else:
            print(f"⚠️ No comparable data in sheet: {sheet_name}")

    wb_out.save(output_path)
    print(f"\n📊 Comparison report saved at: {output_path}")

def compare_widget_data_preserve_structure(excel_path, db_data, output_path, widget_sp_map):
    """
    Enhanced comparison that preserves Excel column structure instead of converting to rows
    """
    def safe_round(value):
        try:
            clean_val = str(value).replace(",", "").replace("$", "").replace("₹", "").replace("K", "").replace("M", "")
            return round(float(clean_val), 2)
        except:
            return value

    def normalize(name: str) -> str:
        return ''.join(str(name).lower().strip().replace("_", "").replace(" ", ""))

    if not os.path.exists(excel_path):
        print("❌ Excel file not found.")
        return

    wb_source = load_workbook(excel_path, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in wb_source.sheetnames:
        print(f"\n📄 Comparing sheet: {sheet_name}")
        ws = wb_source[sheet_name]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        col_indices = {h: i for i, h in enumerate(headers)}

        # Create output sheet with same structure as input
        ws_out = wb_out.create_sheet(title=sheet_name[:31])
        
        # Determine the structure type
        has_targets = "targets" in headers or "target" in headers
        has_prev_curr = "previous year" in headers and "current year" in headers
        has_metric_value = "metric" in headers and "excel value" in headers
        
        if has_targets:
            # Structure: Identifier | Excel Actual | DB Actual | Actual Status | Excel Target | DB Target | Target Status
            ws_out.append(["Identifier", "Excel Actual Sales", "DB Actual Sales", "Actual Sales Status", "Excel Targets", "DB Targets", "Target Status"])
            
            id_idx = col_indices.get("identifier", 0)
            actual_idx = col_indices.get("actual sales")
            target_idx = col_indices.get("targets") or col_indices.get("target")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                identifier = str(row[id_idx]).strip() if row[id_idx] else "Unknown"
                
                # Get Excel values
                excel_actual = row[actual_idx] if actual_idx is not None and len(row) > actual_idx else None
                excel_target = row[target_idx] if target_idx is not None and len(row) > target_idx else None
                
                # Get DB values
                actual_key = f"{sheet_name} - {identifier}"
                target_key = f"{sheet_name} - {identifier} Target"
                
                db_actual = db_data.get(actual_key, "Not Found")
                db_target = db_data.get(target_key, "Not Found")
                
                # Compare values
                actual_status = "Match"
                target_status = "Match"
                
                if excel_actual is not None:
                    excel_actual_rounded = safe_round(excel_actual)
                    db_actual_rounded = safe_round(db_actual) if db_actual != "Not Found" else "Not Found"
                    actual_status = "Match" if excel_actual_rounded == db_actual_rounded else "Mismatch"
                    if db_actual == "Not Found":
                        actual_status = "Not Found"
                
                if excel_target is not None:
                    excel_target_rounded = safe_round(excel_target)
                    db_target_rounded = safe_round(db_target) if db_target != "Not Found" else "Not Found"
                    target_status = "Match" if excel_target_rounded == db_target_rounded else "Mismatch"
                    if db_target == "Not Found":
                        target_status = "Not Found"
                
                ws_out.append([
                    identifier,
                    excel_actual,
                    db_actual,
                    actual_status,
                    excel_target,
                    db_target,
                    target_status
                ])
        
        elif has_prev_curr:
            # Structure: Identifier | Excel Previous | DB Previous | Previous Status | Excel Current | DB Current | Current Status
            ws_out.append(["Identifier", "Excel Previous Year", "DB Previous Year", "Previous Year Status", "Excel Current Year", "DB Current Year", "Current Year Status"])
            
            id_idx = col_indices.get("identifier", 0)
            prev_idx = col_indices.get("previous year")
            curr_idx = col_indices.get("current year")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                identifier = str(row[id_idx]).strip() if row[id_idx] else "Unknown"
                
                # Get Excel values
                excel_prev = row[prev_idx] if prev_idx is not None and len(row) > prev_idx else None
                excel_curr = row[curr_idx] if curr_idx is not None and len(row) > curr_idx else None
                
                # Get DB values
                prev_key = f"{sheet_name} - {identifier} Previous Year"
                curr_key = f"{sheet_name} - {identifier} Current Year"
                
                db_prev = db_data.get(prev_key, "Not Found")
                db_curr = db_data.get(curr_key, "Not Found")
                
                # Compare values
                prev_status = "Match"
                curr_status = "Match"
                
                if excel_prev is not None:
                    excel_prev_rounded = safe_round(excel_prev)
                    db_prev_rounded = safe_round(db_prev) if db_prev != "Not Found" else "Not Found"
                    prev_status = "Match" if excel_prev_rounded == db_prev_rounded else "Mismatch"
                    if db_prev == "Not Found":
                        prev_status = "Not Found"
                
                if excel_curr is not None:
                    excel_curr_rounded = safe_round(excel_curr)
                    db_curr_rounded = safe_round(db_curr) if db_curr != "Not Found" else "Not Found"
                    curr_status = "Match" if excel_curr_rounded == db_curr_rounded else "Mismatch"
                    if db_curr == "Not Found":
                        curr_status = "Not Found"
                
                ws_out.append([
                    identifier,
                    excel_prev,
                    db_prev,
                    prev_status,
                    excel_curr,
                    db_curr,
                    curr_status
                ])
        
        else:
            # Standard structure: Identifier | Excel Actual Sales | DB Actual Sales | Status
            ws_out.append(["Identifier", "Excel Actual Sales", "DB Actual Sales", "Status"])
            
            id_idx = col_indices.get("identifier", 0)
            actual_idx = col_indices.get("actual sales")
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                identifier = str(row[id_idx]).strip() if row[id_idx] else "Unknown"
                
                # Get Excel value
                excel_actual = row[actual_idx] if actual_idx is not None and len(row) > actual_idx else None
                
                # Get DB value
                actual_key = f"{sheet_name} - {identifier}"
                db_actual = db_data.get(actual_key, "Not Found")
                
                # Compare values
                status = "Match"
                if excel_actual is not None:
                    excel_actual_rounded = safe_round(excel_actual)
                    db_actual_rounded = safe_round(db_actual) if db_actual != "Not Found" else "Not Found"
                    status = "Match" if excel_actual_rounded == db_actual_rounded else "Mismatch"
                    if db_actual == "Not Found":
                        status = "Not Found"
                
                ws_out.append([
                    identifier,
                    excel_actual,
                    db_actual,
                    status
                ])

    wb_out.save(output_path)
    print(f"\n📊 Enhanced comparison report saved at: {output_path}")

def dynamic_compare_widget_data(excel_path, output_path, params=None):
    """
    Enhanced widget comparison using dynamic engine - NO HARDCODED MAPPINGS!
    Falls back to legacy method if dynamic engine is not available
    """
    if not DYNAMIC_ENGINE_AVAILABLE:
        print("⚠️ Dynamic engine not available, using legacy comparison")
        return compare_widget_data(excel_path, {}, output_path, widget_sp_map)
    
    try:
        print(f"🧠 Starting dynamic widget comparison...")
        
        # Use default parameters if not provided
        if params is None:
            params = (2024, None, None, None, None, None, None)
        
        # Initialize dynamic engine
        dynamic_engine = DynamicComparisonEngine()
        
        # Get mappings and fetch data
        mappings = dynamic_engine.create_intelligent_mapping(excel_path, params)
        if not mappings:
            print("❌ No valid mappings created!")
            return False
        
        # Fetch DB data
        db_data = dynamic_engine.fetch_dynamic_db_data(mappings, params)
        if not db_data:
            print("❌ No DB data fetched!")
            return False
        
        # Use enhanced comparison that preserves structure
        compare_widget_data_preserve_structure(excel_path, db_data, output_path, mappings)
        
        print(f"✅ Dynamic widget comparison completed successfully!")
        return True
            
    except Exception as e:
        print(f"❌ Dynamic widget comparison error: {str(e)}")
        print(f"🔄 Falling back to legacy method...")
        # Fallback to legacy method
        try:
            if params is None:
                params = (2024, None, None, None, None, None, None)
            db_values = fetch_db_widget_values(widget_sp_map, params)
            compare_widget_data(excel_path, db_values, output_path, widget_sp_map)
            return True
        except Exception as fallback_error:
            print(f"❌ Legacy fallback also failed: {str(fallback_error)}")
            return False

    def safe_round(value):
        try:
            clean_val = str(value).replace(",", "").replace("$", "").replace("₹", "").replace("K", "").replace("M", "")
            return round(float(clean_val), 2)
        except:
            return value

    def normalize(name: str) -> str:
        return ''.join(str(name).lower().strip().replace("_", "").replace(" ", ""))

    if not os.path.exists(excel_path):
        print("❌ Excel file not found.")
        return

    # Normalize widget name mapping
    normalized_widget_map = {normalize(v): v for v in widget_sp_map.values()}
    reverse_widget_map = {normalize(k): k for k in widget_sp_map.keys()}

    wb_source = load_workbook(excel_path, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    for sheet_name in wb_source.sheetnames:
        print(f"\n📄 Comparing sheet: {sheet_name}")
        ws = wb_source[sheet_name]
        headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
        col_indices = {h: i for i, h in enumerate(headers)}

        # Flexible column lookup
        id_idx = col_indices.get("identifier") or col_indices.get("identifier/label") or 0
        actual_idx = col_indices.get("actual sales")
        target_idx = col_indices.get("targets") or col_indices.get("target")
        prev_idx = col_indices.get("previous year")
        curr_idx = col_indices.get("current year")
        metric_idx = col_indices.get("metric")
        excelval_idx = col_indices.get("excel value")

        output = []

        # Resolve widget name
        normalized_sheet = normalize(sheet_name)
        widget_name = None
        for norm_key, display_key in reverse_widget_map.items():
            if norm_key in normalized_sheet:
                widget_name = display_key
                break

        if not widget_name:
            print(f"⚠️ Widget name not mapped for sheet '{sheet_name}', using sheet name as fallback.")
            widget_name = sheet_name

        for row in ws.iter_rows(min_row=2, values_only=True):
            identifier = str(row[id_idx]).strip() if row[id_idx] else "Unknown"

            def compare_metric(label, col_index, suffix=""):
                if col_index is not None and row[col_index] is not None:
                    excel_val = row[col_index]
                    # Use old code's key format for Actual Sales, align with database
                    key = f"{widget_name} - {identifier}" if label == "Actual Sales" else f"{widget_name} - {identifier} {suffix}"
                    db_val = db_data.get(key, "Not Found")
                    excel_rounded = safe_round(excel_val)
                    db_rounded = safe_round(db_val)
                    status = "Match" if excel_rounded == db_rounded else "Mismatch"
                    if db_val == "Not Found":
                        print(f"❌ DB value not found for key: {key}")
                    output.append([identifier, label, excel_rounded, db_rounded, status])

            # Compare fixed columns
            compare_metric("Actual Sales", actual_idx)
            compare_metric("Target", target_idx, "Target")
            compare_metric("Previous Year", prev_idx, "Previous Year")
            compare_metric("Current Year", curr_idx, "Current Year")

            # Compare custom metric-based values
            if metric_idx is not None and excelval_idx is not None:
                metric = str(row[metric_idx]).strip() if row[metric_idx] else ""
                excel_val = row[excelval_idx]
                key = f"{widget_name} - {identifier} {metric}"
                db_val = db_data.get(key, "Not Found")
                excel_rounded = safe_round(excel_val)
                db_rounded = safe_round(db_val)
                status = "Match" if excel_rounded == db_rounded else "Mismatch"
                if db_val == "Not Found":
                    print(f"❌ DB value not found for key: {key}")
                output.append([identifier, metric, excel_rounded, db_rounded, status])

        if output:
            ws_out = wb_out.create_sheet(title=sheet_name[:31])
            ws_out.append(["Identifier/Label", "Metric", "Excel Value", "DB Value", "Status"])
            for line in output:
                ws_out.append(line)
            print(f"✅ Sheet '{sheet_name}' compared with {len(output)} rows.")
        else:
            print(f"⚠️ No comparable data in sheet: {sheet_name}")

    wb_out.save(output_path)
    print(f"\n📊 Comparison report saved at: {output_path}")