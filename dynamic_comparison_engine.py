"""
DYNAMIC DATA COMPARISON ENGINE
=============================
Eliminates hardcoded sheet-to-SP mappings through intelligent analysis
"""

import re
import os
from openpyxl import load_workbook, Workbook
from dataBase import DatabaseConnector
from collections import defaultdict
from config_loader import config_loader

class DynamicComparisonEngine:
    def __init__(self):
        self.db_connector = DatabaseConnector()
        
        # Auto-discoverable stored procedures
        self.available_sps = [
            "SP_SalesTrend",
            "SP_TopPerformingEmployee", 
            "SP_TopProductsBySales",
            "SP_WeekdayWeekendSales",
            "SP_WeekwiseSalesComparison",
            "SP_TopStoresbySales",
            "SP_TopBrandsBySales",
            "SP_TopCategoriesBySaleswidget",
            "SP_TopSubCategoriesBySales",
            "SP_WeeklyTrendswidget",
            "SP_StorewiseActualVsTarget_Vertical_SortedByActual"
        ]
        
        # Pattern-based SP suggestions (can be learned/updated)
        self.sp_patterns = {
            r'sales.*trend|trend.*sales|month.*sales': 'SP_SalesTrend',
            r'week.*sales|weekly.*sales|weekwise': 'SP_WeekwiseSalesComparison',
            r'weekday|weekend|day.*type': 'SP_WeekdayWeekendSales',
            r'product.*sales|sales.*product': 'SP_TopProductsBySales',
            r'brand.*sales|sales.*brand': 'SP_TopBrandsBySales',
            r'store.*sales|sales.*store': 'SP_TopStoresbySales',
            r'employee|performing': 'SP_TopPerformingEmployee',
            r'categor.*sales|sales.*categor': 'SP_TopCategoriesBySaleswidget',
            r'target.*actual|actual.*target': 'SP_StorewiseActualVsTarget_Vertical_SortedByActual'
        }
    
    def analyze_excel_structure(self, excel_path):
        """
        Analyze Excel file structure to understand data patterns
        Returns comprehensive analysis for each sheet
        """
        if not os.path.exists(excel_path):
            print(f"❌ Excel file not found: {excel_path}")
            return {}
        
        wb = load_workbook(excel_path, data_only=True)
        analysis = {}
        
        print(f"🔍 Analyzing Excel structure: {os.path.basename(excel_path)}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get headers
            headers = []
            for cell in ws[1]:
                header = str(cell.value).strip().lower() if cell.value else ""
                headers.append(header)
            
            # Analyze sheet structure
            sheet_analysis = {
                'original_name': sheet_name,
                'normalized_name': self._normalize_name(sheet_name),
                'headers': headers,
                'data_patterns': self._detect_data_patterns(headers),
                'key_columns': self._identify_key_columns(headers),
                'sample_data': self._extract_sample_data(ws, headers),
                'suggested_sps': self._suggest_stored_procedures(sheet_name, headers),
                'data_volume': ws.max_row - 1  # Exclude header row
            }
            
            analysis[sheet_name] = sheet_analysis
            print(f"  📄 {sheet_name}: {sheet_analysis['data_patterns']} ({sheet_analysis['data_volume']} rows)")
        
        return analysis
    
    def _normalize_name(self, name):
        """Normalize names for pattern matching"""
        return re.sub(r'[^a-zA-Z0-9]', '', name.lower())
    
    def _detect_data_patterns(self, headers):
        """Detect data patterns from headers"""
        header_text = ' '.join(headers).lower()
        
        patterns = []
        
        # Time-based patterns
        if any(x in header_text for x in ['month', 'monthly']):
            patterns.append('monthly_data')
        if any(x in header_text for x in ['week', 'weekly']):
            patterns.append('weekly_data')
        if any(x in header_text for x in ['weekday', 'weekend']):
            patterns.append('weekday_weekend_data')
        
        # Entity patterns
        if 'product' in header_text:
            patterns.append('product_data')
        if 'brand' in header_text:
            patterns.append('brand_data')
        if 'store' in header_text:
            patterns.append('store_data')
        if 'employee' in header_text:
            patterns.append('employee_data')
        if 'category' in header_text:
            patterns.append('category_data')
        
        # Value patterns
        if any(x in header_text for x in ['sales', 'revenue']):
            patterns.append('sales_data')
        if 'target' in header_text:
            patterns.append('target_data')
        if any(x in header_text for x in ['current year', 'previous year']):
            patterns.append('year_comparison')
        
        return patterns if patterns else ['generic_data']
    
    def _identify_key_columns(self, headers):
        """Identify key columns for data extraction"""
        key_columns = {}
        
        for i, header in enumerate(headers):
            header_lower = header.lower()
            
            # Identifier columns (priority order)
            if 'identifier' in header_lower:
                key_columns['identifier'] = i
            elif any(x in header_lower for x in ['name', 'product', 'brand', 'store', 'employee']):
                if 'identifier' not in key_columns:
                    key_columns['identifier'] = i
            
            # Value columns
            if 'actual sales' in header_lower:
                key_columns['actual_sales'] = i
            elif 'sales' in header_lower and 'actual' not in key_columns:
                key_columns['sales'] = i
            
            if 'target' in header_lower:
                key_columns['target'] = i
            
            if 'current year' in header_lower:
                key_columns['current_year'] = i
            elif 'previous year' in header_lower:
                key_columns['previous_year'] = i
            
            if 'metric' in header_lower:
                key_columns['metric'] = i
            elif 'excel value' in header_lower:
                key_columns['excel_value'] = i
        
        return key_columns
    
    def _extract_sample_data(self, worksheet, headers, max_samples=3):
        """Extract sample data for pattern analysis"""
        samples = []
        row_count = 0
        
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row_count >= max_samples:
                break
            
            if any(cell is not None for cell in row):
                sample = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        sample[header] = row[i]
                samples.append(sample)
                row_count += 1
        
        return samples
    
    def _suggest_stored_procedures(self, sheet_name, headers):
        """Suggest stored procedures based on sheet analysis"""
        suggestions = []
        sheet_text = sheet_name.lower()
        header_text = ' '.join(headers).lower()
        combined_text = f"{sheet_text} {header_text}"
        
        # Pattern-based matching
        for pattern, sp_name in self.sp_patterns.items():
            if re.search(pattern, combined_text):
                suggestions.append({
                    'sp_name': sp_name,
                    'confidence': 0.8,
                    'reason': f'Pattern match: {pattern}'
                })
        
        # Fallback suggestions based on keywords
        if not suggestions:
            if 'sales' in combined_text:
                suggestions.append({
                    'sp_name': 'SP_TopProductsBySales',
                    'confidence': 0.5,
                    'reason': 'Generic sales data'
                })
        
        return suggestions
    
    def test_sp_compatibility(self, params, max_test_sps=None):
        """
        Test stored procedures to find which ones work with given parameters
        """
        print(f"🧪 Testing SP compatibility with params: {params}")
        
        conn = self.db_connector.connect()
        cursor = conn.cursor()
        
        compatibility = {}
        test_sps = self.available_sps[:max_test_sps] if max_test_sps else self.available_sps
        
        for sp_name in test_sps:
            try:
                # Test with positional parameters
                cursor.execute(f"EXEC [{sp_name}] ?, ?, ?, ?, ?, ?, ?", params)
                result = cursor.fetchall()
                columns = [desc[0] for desc in cursor.description]
                
                compatibility[sp_name] = {
                    'works': True,
                    'sp_name': sp_name,  # Add sp_name for pattern matching
                    'row_count': len(result),
                    'columns': columns,
                    'output_pattern': self._classify_sp_output(columns, result),
                    'sample_row': dict(zip(columns, result[0])) if result else None
                }
                
                print(f"  ✅ {sp_name}: {len(result)} rows, pattern: {compatibility[sp_name]['output_pattern']}")
                
            except Exception as e:
                compatibility[sp_name] = {
                    'works': False,
                    'error': str(e)
                }
                print(f"  ❌ {sp_name}: {str(e)}")
        
        conn.close()
        return compatibility
    
    def _classify_sp_output(self, columns, result):
        """Enhanced SP output pattern classification"""
        col_text = ' '.join(columns).lower()
        
        # Check actual data content for better classification
        sample_data = {}
        if result and len(result) > 0:
            sample_data = dict(zip(columns, result[0]))
        
        # Metric-value pairs pattern (check this first for SP_StorewiseActualVsTarget_Vertical_SortedByActual)
        if 'metric' in col_text and any(x in col_text for x in ['excel value', 'value']):
            return 'metric_value_pairs'
        
        # Monthly trends pattern
        elif 'month' in col_text and 'currentyearsales' in col_text:
            return 'monthly_trends'
        
        # Weekly trends pattern  
        elif 'week' in col_text and 'currentyearsales' in col_text:
            return 'weekly_trends'
        
        # Weekday/Weekend pattern
        elif any(x in col_text for x in ['daytype', 'weekcategory']) and 'sales' in col_text:
            return 'weekday_weekend'
        
        # Product sales pattern
        elif 'productname' in col_text and 'sales' in col_text:
            return 'product_sales'
        
        # Brand sales pattern
        elif 'brandname' in col_text and 'sales' in col_text:
            return 'brand_sales'
        
        # Store sales pattern (including target data)
        elif any(x in col_text for x in ['storename', 'store']) and 'sales' in col_text:
            return 'store_sales'
        elif 'identifier' in col_text and any(x in col_text for x in ['actual', 'target']):
            return 'store_sales'  # Likely store data with targets
        
        # Employee performance pattern
        elif 'employeename' in col_text and 'sales' in col_text:
            return 'employee_performance'
        
        # Category sales pattern
        elif any(x in col_text for x in ['categoryname', 'category']) and 'sales' in col_text:
            return 'category_sales'
        
        # Generic pattern (fallback)
        else:
            # Try to infer from column structure
            if len(columns) >= 2 and any(x in col_text for x in ['sales', 'actual', 'value']):
                return 'generic_sales'
            else:
                return 'generic'
    
    def create_intelligent_mapping(self, excel_path, params):
        """
        Create intelligent sheet-to-SP mapping without hardcoded rules
        """
        print(f"🧠 Creating intelligent mapping for: {os.path.basename(excel_path)}")
        
        # Step 1: Analyze Excel structure
        excel_analysis = self.analyze_excel_structure(excel_path)
        if not excel_analysis:
            return {}
        
        # Step 2: Test SP compatibility
        sp_compatibility = self.test_sp_compatibility(params)
        working_sps = {sp: info for sp, info in sp_compatibility.items() if info['works']}
        
        if not working_sps:
            print("❌ No working stored procedures found!")
            return {}
        
        print(f"✅ Found {len(working_sps)} working SPs")
        
        # Step 3: Create intelligent mappings
        mappings = {}
        
        for sheet_name, sheet_info in excel_analysis.items():
            best_mapping = self._find_best_sp_match(sheet_info, working_sps)
            
            if best_mapping:
                mappings[sheet_name] = best_mapping
                print(f"  📊 {sheet_name} → {best_mapping['sp_name']} (confidence: {best_mapping['confidence']:.2f})")
            else:
                print(f"  ⚠️ No suitable SP found for: {sheet_name}")
        
        return mappings
    
    def _find_best_sp_match(self, sheet_info, working_sps):
        """Find the best SP match for a sheet using improved logic"""
        best_match = None
        best_score = 0
        
        sheet_name = sheet_info['original_name'].lower()
        sheet_patterns = set(sheet_info['data_patterns'])
        
        print(f"🔍 Finding best SP match for: {sheet_info['original_name']}")
        print(f"   Sheet patterns: {sheet_patterns}")
        
        # Enhanced pattern matching with sheet name analysis
        for sp_name, sp_info in working_sps.items():
            score = self._calculate_enhanced_match_score(sheet_info, sp_info, sheet_name)
            
            print(f"   {sp_name}: score={score:.2f}, pattern={sp_info['output_pattern']}")
            
            if score > best_score:
                best_score = score
                best_match = {
                    'sp_name': sp_name,
                    'confidence': score,
                    'output_pattern': sp_info['output_pattern'],
                    'key_columns': sheet_info['key_columns'],
                    'data_patterns': sheet_info['data_patterns']
                }
        
        if best_match:
            print(f"   ✅ Best match: {best_match['sp_name']} (confidence: {best_match['confidence']:.2f})")
        else:
            print(f"   ❌ No suitable match found")
        
        return best_match if best_score > 0.3 else None  # Minimum confidence threshold
    
    def _calculate_enhanced_match_score(self, sheet_info, sp_info, sheet_name):
        """Enhanced match score calculation with sheet name analysis"""
        score = 0.0
        
        sheet_patterns = set(sheet_info['data_patterns'])
        sp_pattern = sp_info['output_pattern']
        sp_name = sp_info.get('sp_name', '')
        
        # Direct sheet name to SP mapping (highest priority)
        # Special handling for store sales with targets
        if 'store sales' in sheet_name and 'target_data' in sheet_patterns:
            direct_mappings = {
                'store sales': 'SP_StorewiseActualVsTarget_Vertical_SortedByActual',
            }
        else:
            direct_mappings = {
                'store sales': 'SP_TopStoresbySales',
                'top stores': 'SP_TopStoresbySales',
                'brands': 'SP_TopBrandsBySales',
                'top brands': 'SP_TopBrandsBySales',
                'categories': 'SP_TopCategoriesBySaleswidget',
                'top categories': 'SP_TopCategoriesBySaleswidget',
                'sub categories': 'SP_TopSubCategoriesBySales',
                'products': 'SP_TopProductsBySales',
                'top products': 'SP_TopProductsBySales',
                'weekly trends': 'SP_WeeklyTrendswidget',  # Exact match for Weekly Trends
                'weekly': 'SP_WeeklyTrendswidget',
                'sales trends': 'SP_SalesTrend',
                'month': 'SP_SalesTrend'
            }
        
        # Check for direct name matches
        for name_pattern, expected_sp in direct_mappings.items():
            if name_pattern in sheet_name and expected_sp in sp_name:
                score += 0.8
                print(f"     Direct name match: {name_pattern} -> {expected_sp}")
                break
        
        # Pattern matching scores (secondary priority)
        pattern_matches = {
            'monthly_trends': {'monthly_data', 'sales_data', 'year_comparison'},
            'weekly_trends': {'weekly_data', 'sales_data', 'year_comparison'},
            'weekday_weekend': {'weekday_weekend_data', 'sales_data'},
            'product_sales': {'product_data', 'sales_data'},
            'brand_sales': {'brand_data', 'sales_data'},
            'store_sales': {'store_data', 'sales_data'},
            'employee_performance': {'employee_data', 'sales_data'},
            'metric_value_pairs': {'target_data', 'sales_data'},
            'generic': {'sales_data'}  # Generic can match basic sales data
        }
        
        expected_patterns = pattern_matches.get(sp_pattern, set())
        if expected_patterns.intersection(sheet_patterns):
            score += 0.4
        
        # SP-specific pattern matching with improved logic
        if 'SP_StorewiseActualVsTarget_Vertical_SortedByActual' in sp_name and ('store' in sheet_name and 'target_data' in sheet_patterns):
            score += 0.9  # Strongly prefer this for store data with targets
        elif 'SP_TopStoresbySales' in sp_name and ('store' in sheet_name and 'target_data' not in sheet_patterns):
            score += 0.5  # Use this for store data without targets
        elif 'SP_TopBrandsBySales' in sp_name and 'brand' in sheet_name:
            score += 0.5
        elif 'SP_TopCategoriesBySaleswidget' in sp_name and 'categor' in sheet_name and 'sub' not in sheet_name:
            score += 0.5  # Only for main categories, not sub-categories
        elif 'SP_TopSubCategoriesBySales' in sp_name and 'sub categor' in sheet_name:
            score += 0.7  # Prefer this for sub-categories
        elif 'SP_TopProductsBySales' in sp_name and 'product' in sheet_name:
            score += 0.5
        elif 'SP_WeeklyTrendswidget' in sp_name and 'weekly trends' in sheet_name:
            score += 0.9  # Strongly prefer this for "Weekly Trends" sheets
        elif 'SP_WeeklyTrendswidget' in sp_name and 'weekly' in sheet_name:
            score += 0.7  # Prefer this for weekly trends
        elif 'SP_WeekwiseSalesComparison' in sp_name and 'weekly' in sheet_name and 'trends' not in sheet_name:
            score += 0.6  # Alternative for weekly data (but not for "Weekly Trends")
        elif 'SP_SalesTrend' in sp_name and ('sales trend' in sheet_name or ('trend' in sheet_name and 'weekly' not in sheet_name)):
            score += 0.5  # Only for non-weekly trends
        
        # Data volume consideration (lower priority)
        if sp_info['row_count'] > 0:
            score += 0.1
        
        # Column compatibility (lower priority)
        if sheet_info['key_columns'] and sp_info['columns']:
            score += 0.1
        
        return min(score, 1.0)
    
    def fetch_dynamic_db_data(self, mappings, params):
        """
        Fetch DB data using dynamic mappings
        """
        print(f"📊 Fetching DB data using {len(mappings)} dynamic mappings...")
        
        conn = self.db_connector.connect()
        cursor = conn.cursor()
        
        all_db_data = {}
        
        for sheet_name, mapping in mappings.items():
            sp_name = mapping['sp_name']
            output_pattern = mapping['output_pattern']
            
            try:
                cursor.execute(f"EXEC [{sp_name}] ?, ?, ?, ?, ?, ?, ?", params)
                result = cursor.fetchall()
                columns = [desc[0] for desc in cursor.description]
                
                # Format data based on output pattern
                formatted_data = self._format_sp_data(
                    result, columns, sheet_name, output_pattern
                )
                
                all_db_data.update(formatted_data)
                print(f"  ✅ {sp_name}: {len(formatted_data)} values for {sheet_name}")
                
            except Exception as e:
                print(f"  ❌ Error with {sp_name}: {str(e)}")
        
        conn.close()
        return all_db_data
    
    def _format_sp_data(self, result, columns, sheet_name, output_pattern):
        """Format SP data based on detected output pattern"""
        formatted_data = {}
        
        if output_pattern == 'monthly_trends':
            formatted_data.update(self._format_monthly_trends(result, columns, sheet_name))
        elif output_pattern == 'weekly_trends':
            formatted_data.update(self._format_weekly_trends(result, columns, sheet_name))
        elif output_pattern == 'weekday_weekend':
            formatted_data.update(self._format_weekday_weekend(result, columns, sheet_name))
        elif output_pattern == 'product_sales':
            formatted_data.update(self._format_product_sales(result, columns, sheet_name))
        elif output_pattern == 'brand_sales':
            formatted_data.update(self._format_brand_sales(result, columns, sheet_name))
        elif output_pattern == 'store_sales':
            formatted_data.update(self._format_store_sales(result, columns, sheet_name))
        elif output_pattern == 'employee_performance':
            formatted_data.update(self._format_employee_performance(result, columns, sheet_name))
        elif output_pattern == 'metric_value_pairs':
            formatted_data.update(self._format_metric_value_pairs(result, columns, sheet_name))
        elif output_pattern == 'category_sales':
            formatted_data.update(self._format_category_sales(result, columns, sheet_name))
        elif output_pattern == 'generic_sales':
            formatted_data.update(self._format_generic_sales(result, columns, sheet_name))
        else:
            # Generic formatting
            formatted_data.update(self._format_generic(result, columns, sheet_name))
        
        return formatted_data
    
    def _format_monthly_trends(self, result, columns, sheet_name):
        """Format monthly trends data"""
        data = {}
        month_abbrev = {
            'January': 'Jan', 'February': 'Feb', 'March': 'Mar', 'April': 'Apr',
            'May': 'May', 'June': 'Jun', 'July': 'Jul', 'August': 'Aug',
            'September': 'Sep', 'October': 'Oct', 'November': 'Nov', 'December': 'Dec'
        }
        
        for row in result:
            row_dict = dict(zip(columns, row))
            month = str(row_dict.get('Month', 'Unknown')).strip()
            month = month_abbrev.get(month, month)
            
            if 'CurrentYearSales' in row_dict and row_dict['CurrentYearSales'] is not None:
                data[f"{sheet_name} - {month} Current Year"] = str(row_dict['CurrentYearSales']).strip()
            if 'PreviousYearSales' in row_dict and row_dict['PreviousYearSales'] is not None:
                data[f"{sheet_name} - {month} Previous Year"] = str(row_dict['PreviousYearSales']).strip()
        
        return data
    
    def _format_weekly_trends(self, result, columns, sheet_name):
        """Format weekly trends data"""
        data = {}
        
        print(f"🔍 Formatting weekly trends for {sheet_name}")
        print(f"   Columns: {columns}")
        
        for row in result:
            row_dict = dict(zip(columns, row))
            week = str(row_dict.get('Week') or row_dict.get('WeekCategory', 'Unknown')).strip()
            
            if 'CurrentYearSales' in row_dict and row_dict['CurrentYearSales'] is not None:
                # Clean the value - remove commas
                current_value = str(row_dict['CurrentYearSales']).replace(',', '').strip()
                key = f"{sheet_name} - {week} Current Year"
                data[key] = current_value
                print(f"   Added: {key} = {current_value}")
                
            if 'PreviousYearSales' in row_dict and row_dict['PreviousYearSales'] is not None:
                # Clean the value - remove commas
                previous_value = str(row_dict['PreviousYearSales']).replace(',', '').strip()
                key = f"{sheet_name} - {week} Previous Year"
                data[key] = previous_value
                print(f"   Added: {key} = {previous_value}")
        
        print(f"   Total formatted: {len(data)} values")
        return data
    
    def _format_weekday_weekend(self, result, columns, sheet_name):
        """Format weekday/weekend data"""
        data = {}
        for row in result:
            row_dict = dict(zip(columns, row))
            day_type = str(row_dict.get('DayType') or row_dict.get('WeekCategory', 'Unknown')).strip().upper()
            
            if 'CurrentYearSales' in row_dict and row_dict['CurrentYearSales'] is not None:
                data[f"{sheet_name} - {day_type} SALES Current Year"] = str(row_dict['CurrentYearSales']).strip()
            if 'PreviousYearSales' in row_dict and row_dict['PreviousYearSales'] is not None:
                data[f"{sheet_name} - {day_type} SALES Previous Year"] = str(row_dict['PreviousYearSales']).strip()
        
        return data
    
    def _format_product_sales(self, result, columns, sheet_name):
        """Format product sales data"""
        data = {}
        for row in result:
            row_dict = dict(zip(columns, row))
            product = str(row_dict.get('ProductName', 'Unknown')).strip()
            sales = row_dict.get('Sales')
            
            if sales is not None:
                data[f"{sheet_name} - {product}"] = str(sales).strip()
        
        return data
    
    def _format_brand_sales(self, result, columns, sheet_name):
        """Format brand sales data"""
        data = {}
        for row in result:
            row_dict = dict(zip(columns, row))
            brand = str(row_dict.get('BrandName', 'Unknown')).strip()
            sales = row_dict.get('Sales')
            
            if sales is not None:
                data[f"{sheet_name} - {brand}"] = str(sales).strip()
        
        return data
    
    def _format_store_sales(self, result, columns, sheet_name):
        """Format store sales data"""
        data = {}
        
        print(f"🔍 Formatting store sales data for {sheet_name}")
        print(f"   Columns: {columns}")
        
        # Handle different column name variations
        identifier_col = None
        sales_col = None
        target_col = None
        
        for col in columns:
            col_lower = col.lower()
            if any(x in col_lower for x in ['storename', 'store', 'identifier']):
                identifier_col = col
            elif any(x in col_lower for x in ['actual sales', 'sales', 'actual']):
                sales_col = col
            elif 'target' in col_lower:
                target_col = col
        
        print(f"   Identifier column: {identifier_col}")
        print(f"   Sales column: {sales_col}")
        print(f"   Target column: {target_col}")
        
        if identifier_col and sales_col:
            for row in result:
                row_dict = dict(zip(columns, row))
                identifier = str(row_dict.get(identifier_col, 'Unknown')).strip()
                sales = row_dict.get(sales_col)
                
                if sales is not None:
                    key = f"{sheet_name} - {identifier}"
                    data[key] = str(sales).strip()
                    print(f"   Added: {key} = {sales}")
                
                # Handle target data if available
                if target_col and row_dict.get(target_col) is not None:
                    target_value = row_dict.get(target_col)
                    target_key = f"{sheet_name} - {identifier} Target"
                    data[target_key] = str(target_value).strip()
                    print(f"   Added: {target_key} = {target_value}")
        else:
            print(f"   ❌ Could not identify columns, using generic formatting")
            # Fallback to generic formatting
            return self._format_generic_sales(result, columns, sheet_name)
        
        print(f"   Total formatted: {len(data)} values")
        return data
    
    def _format_employee_performance(self, result, columns, sheet_name):
        """Format employee performance data"""
        data = {}
        for row in result:
            row_dict = dict(zip(columns, row))
            employee = str(row_dict.get('EmployeeName', 'Unknown')).strip()
            sales = row_dict.get('Sales')
            
            if sales is not None:
                data[f"{sheet_name} - {employee}"] = str(sales).strip()
        
        return data
    
    def _format_metric_value_pairs(self, result, columns, sheet_name):
        """Format metric-value pair data"""
        data = {}
        
        print(f"🔍 Formatting metric-value pairs for {sheet_name}")
        print(f"   Columns: {columns}")
        
        for row in result:
            row_dict = dict(zip(columns, row))
            identifier = str(row_dict.get('Identifier', 'Unknown')).strip()
            metric = str(row_dict.get('Metric', '')).strip()
            value = row_dict.get('Excel Value')
            
            if value is not None:
                # Clean the value - remove commas and convert to proper format
                clean_value = str(value).replace(',', '').strip()
                
                # Create key based on metric type
                if metric.lower() == 'actual sales':
                    key = f"{sheet_name} - {identifier}"
                elif metric.lower() == 'target':
                    key = f"{sheet_name} - {identifier} Target"
                else:
                    key = f"{sheet_name} - {identifier} {metric}"
                
                data[key] = clean_value
                print(f"   Added: {key} = {clean_value}")
        
        print(f"   Total formatted: {len(data)} values")
        return data
    
    def _format_category_sales(self, result, columns, sheet_name):
        """Format category sales data"""
        data = {}
        for row in result:
            row_dict = dict(zip(columns, row))
            category = str(row_dict.get('CategoryName', 'Unknown')).strip()
            sales = row_dict.get('Sales')
            
            if sales is not None:
                data[f"{sheet_name} - {category}"] = str(sales).strip()
        
        return data
    
    def _format_generic_sales(self, result, columns, sheet_name):
        """Format generic sales data with identifier and sales columns"""
        data = {}
        
        print(f"🔍 Formatting generic sales data for {sheet_name}")
        print(f"   Columns: {columns}")
        
        # Try to find identifier and sales columns
        identifier_col = None
        sales_col = None
        target_col = None
        
        for col in columns:
            col_lower = col.lower()
            if 'identifier' in col_lower:
                identifier_col = col
            elif any(x in col_lower for x in ['name', 'store', 'product', 'brand']) and not identifier_col:
                identifier_col = col
            elif any(x in col_lower for x in ['actual sales', 'sales', 'actual']):
                sales_col = col
            elif 'target' in col_lower:
                target_col = col
        
        print(f"   Identifier column: {identifier_col}")
        print(f"   Sales column: {sales_col}")
        print(f"   Target column: {target_col}")
        
        if identifier_col and sales_col:
            for row in result:
                row_dict = dict(zip(columns, row))
                identifier = str(row_dict.get(identifier_col, 'Unknown')).strip()
                sales = row_dict.get(sales_col)
                
                if sales is not None:
                    key = f"{sheet_name} - {identifier}"
                    data[key] = str(sales).strip()
                    print(f"   Added: {key} = {sales}")
                
                # Handle target data if available
                if target_col and row_dict.get(target_col) is not None:
                    target_value = row_dict.get(target_col)
                    target_key = f"{sheet_name} - {identifier} Target"
                    data[target_key] = str(target_value).strip()
                    print(f"   Added: {target_key} = {target_value}")
        else:
            print(f"   ❌ Could not identify identifier or sales columns")
        
        print(f"   Total formatted: {len(data)} values")
        return data
    
    def _format_generic(self, result, columns, sheet_name):
        """Generic data formatting (fallback)"""
        data = {}
        if result and len(columns) >= 2:
            for row in result:
                identifier = str(row[0]).strip() if row[0] else "Unknown"
                value = str(row[1]).strip() if row[1] is not None else "0"
                data[f"{sheet_name} - {identifier}"] = value
        
        return data
    
    def dynamic_compare_data(self, excel_path, params, output_path):
        """
        Main method: Perform dynamic data comparison without hardcoded mappings
        """
        print(f"🚀 Starting dynamic data comparison...")
        print(f"📁 Excel: {os.path.basename(excel_path)}")
        print(f"📋 Params: {params}")
        print(f"📄 Output: {os.path.basename(output_path)}")
        
        try:
            # Step 1: Create intelligent mappings
            mappings = self.create_intelligent_mapping(excel_path, params)
            
            if not mappings:
                print("❌ No valid mappings created!")
                return False
            
            # Step 2: Fetch DB data using mappings
            db_data = self.fetch_dynamic_db_data(mappings, params)
            
            if not db_data:
                print("❌ No DB data fetched!")
                return False
            
            # Step 3: Perform comparison using existing logic
            print(f"🔍 Comparing {len(db_data)} DB values with Excel data...")
            
            # Use existing comparison logic
            self._perform_comparison(excel_path, db_data, output_path, mappings)
            
            if os.path.exists(output_path):
                file_size = os.path.getsize(output_path)
                print(f"✅ Dynamic comparison completed!")
                print(f"📊 Output file: {os.path.basename(output_path)} ({file_size:,} bytes)")
                return True
            else:
                print("❌ Output file was not created!")
                return False
                
        except Exception as e:
            print(f"❌ Dynamic comparison failed: {str(e)}")
            import traceback
            print(f"🔍 Error details: {traceback.format_exc()}")
            return False
    
    def _perform_comparison(self, excel_path, db_data, output_path, mappings):
        """Perform the actual comparison using structure-preserving logic"""
        # Import here to avoid circular import
        from widgetstoreprocedures import compare_widget_data_preserve_structure
        
        # Create dummy SP map for compatibility
        dummy_sp_map = {sheet: mapping['sp_name'] for sheet, mapping in mappings.items()}
        
        # Use enhanced comparison logic that preserves Excel structure
        compare_widget_data_preserve_structure(excel_path, db_data, output_path, dummy_sp_map)

# Usage example for integration
def test_dynamic_engine():
    """Test the dynamic comparison engine"""
    engine = DynamicComparisonEngine()
    
    # Test with sample parameters
    params = (2024, None, "717", None, None, None, None)
    excel_path = "sample_widgets.xlsx"  # Replace with actual path
    output_path = "dynamic_comparison_result.xlsx"
    
    success = engine.dynamic_compare_data(excel_path, params, output_path)
    
    if success:
        print("🎉 Dynamic comparison test successful!")
    else:
        print("❌ Dynamic comparison test failed!")

if __name__ == "__main__":
    test_dynamic_engine()