import os
import time
import logging
from openpyxl import Workbook

class ErrorHandler:
    """Centralized error handling and reporting"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.error_dir = "errors"
        os.makedirs(self.error_dir, exist_ok=True)
    
    def log_error(self, component, widget_name, error_type, error_message, screenshot_path=None):
        """Log an error with details"""
        error_entry = {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "component": component,
            "widget_name": widget_name,
            "error_type": error_type,
            "error_message": str(error_message),
            "screenshot_path": screenshot_path
        }
        self.errors.append(error_entry)
        
        # Also log to console and file
        error_msg = f"❌ {component} - {widget_name}: {error_type} - {error_message}"
        print(error_msg)
        logging.error(error_msg)
    
    def log_warning(self, component, widget_name, warning_type, warning_message):
        """Log a warning with details"""
        warning_entry = {
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "component": component,
            "widget_name": widget_name,
            "warning_type": warning_type,
            "warning_message": str(warning_message)
        }
        self.warnings.append(warning_entry)
        
        # Also log to console and file
        warning_msg = f"⚠️ {component} - {widget_name}: {warning_type} - {warning_message}"
        print(warning_msg)
        logging.warning(warning_msg)
    
    def take_screenshot(self, driver, widget_name, error_type):
        """Take screenshot for error debugging"""
        try:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            safe_widget_name = widget_name.replace(" ", "_").replace("/", "_")
            screenshot_filename = f"{safe_widget_name}_{error_type}_{timestamp}.png"
            screenshot_path = os.path.join(self.error_dir, screenshot_filename)
            
            driver.save_screenshot(screenshot_path)
            print(f"📸 Screenshot saved: {screenshot_path}")
            return screenshot_path
        except Exception as e:
            print(f"⚠️ Could not save screenshot: {str(e)}")
            return None
    
    def generate_error_report(self, output_path="errors/error_report.xlsx"):
        """Generate comprehensive error report"""
        try:
            wb = Workbook()
            
            # Error sheet
            if self.errors:
                ws_errors = wb.active
                ws_errors.title = "Errors"
                ws_errors.append(["Timestamp", "Component", "Widget Name", "Error Type", "Error Message", "Screenshot Path"])
                
                for error in self.errors:
                    ws_errors.append([
                        error["timestamp"],
                        error["component"],
                        error["widget_name"],
                        error["error_type"],
                        error["error_message"],
                        error.get("screenshot_path", "")
                    ])
            
            # Warning sheet
            if self.warnings:
                ws_warnings = wb.create_sheet("Warnings")
                ws_warnings.append(["Timestamp", "Component", "Widget Name", "Warning Type", "Warning Message"])
                
                for warning in self.warnings:
                    ws_warnings.append([
                        warning["timestamp"],
                        warning["component"],
                        warning["widget_name"],
                        warning["warning_type"],
                        warning["warning_message"]
                    ])
            
            # Summary sheet
            ws_summary = wb.create_sheet("Summary")
            ws_summary.append(["Metric", "Count"])
            ws_summary.append(["Total Errors", len(self.errors)])
            ws_summary.append(["Total Warnings", len(self.warnings)])
            
            # Error breakdown by component
            error_by_component = {}
            for error in self.errors:
                component = error["component"]
                error_by_component[component] = error_by_component.get(component, 0) + 1
            
            ws_summary.append([])
            ws_summary.append(["Errors by Component", ""])
            for component, count in error_by_component.items():
                ws_summary.append([component, count])
            
            # Warning breakdown by component
            warning_by_component = {}
            for warning in self.warnings:
                component = warning["component"]
                warning_by_component[component] = warning_by_component.get(component, 0) + 1
            
            ws_summary.append([])
            ws_summary.append(["Warnings by Component", ""])
            for component, count in warning_by_component.items():
                ws_summary.append([component, count])
            
            # Save report
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            wb.save(output_path)
            
            print(f"📊 Error report generated: {output_path}")
            print(f"📈 Summary: {len(self.errors)} errors, {len(self.warnings)} warnings")
            
        except Exception as e:
            print(f"❌ Failed to generate error report: {str(e)}")
    
    def print_summary(self):
        """Print error and warning summary to console"""
        print("\n" + "="*50)
        print("📊 AUTOMATION SUMMARY")
        print("="*50)
        print(f"Total Errors: {len(self.errors)}")
        print(f"Total Warnings: {len(self.warnings)}")
        
        if self.errors:
            print("\n❌ ERRORS:")
            for error in self.errors[-5:]:  # Show last 5 errors
                print(f"  • {error['component']} - {error['widget_name']}: {error['error_type']}")
        
        if self.warnings:
            print("\n⚠️ WARNINGS:")
            for warning in self.warnings[-5:]:  # Show last 5 warnings
                print(f"  • {warning['component']} - {warning['widget_name']}: {warning['warning_type']}")
        
        print("="*50)

# Global error handler instance
error_handler = ErrorHandler()