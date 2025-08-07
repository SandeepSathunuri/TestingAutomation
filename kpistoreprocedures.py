from openpyxl import load_workbook, Workbook
import os
from dataBase import DatabaseConnector
from dotenv import load_dotenv
from config_loader import config_loader

load_dotenv()

databaseconnect = DatabaseConnector()
conn = databaseconnect.connect()
cursor = conn.cursor()

# ✅ Load KPI configurations from YAML
def get_landing_kpi_map():
    """Get landing page KPI stored procedures from YAML config"""
    sp_config = config_loader.get_stored_procedures()
    return sp_config.get("landing_page_kpis", {})

def get_drillthrough_kpi_map():
    """Get drillthrough KPI stored procedures from YAML config"""
    sp_config = config_loader.get_stored_procedures()
    return sp_config.get("drillthrough_kpis", {})

def get_kpi_parameters():
    """Get KPI parameters from YAML config"""
    filter_config = config_loader.get_drillthrough_filters()
    default_params = filter_config.get("default_parameters", {})
    
    return {
        "Year": default_params.get("year", 2024),
        "Month": default_params.get("month"),
        "Store": default_params.get("store"),
        "State": default_params.get("state"),
        "Channel": default_params.get("channel"),
        "FromDate": default_params.get("fromdate"),
        "ToDate": default_params.get("todate")
    }

# 🟦 Landing page stored procedures (Loaded from YAML)
LANDING_SP_MAP = get_landing_kpi_map()

# 🟨 Drillthrough stored procedures (Loaded from YAML)
DRILLTHROUGH_SP_MAP = get_drillthrough_kpi_map()

# ✅ Common parameters for KPIs (Loaded from YAML)
LANDING_PARAMS = get_kpi_parameters()

print(f"📊 Loaded {len(LANDING_SP_MAP)} landing KPI procedures from YAML")
print(f"🎯 Loaded {len(DRILLTHROUGH_SP_MAP)} drillthrough KPI procedures from YAML")

def read_kpi_from_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    kpi_dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
        kpi_name, kpi_value = row
        kpi_dict[kpi_name.strip()] = str(kpi_value).strip()
    return kpi_dict

class Data:
    def __init__(self, is_drillthrough=False, extra_params=None):
        self.cursor = conn.cursor()
        self.is_drillthrough = is_drillthrough
        self.extra_params = extra_params or {}

    def fetch_db_kpi_values(self):
        db_kpi_data = {}

        # Choose map based on drillthrough
        sp_map = DRILLTHROUGH_SP_MAP if self.is_drillthrough else LANDING_SP_MAP
        base_params = LANDING_PARAMS.copy()

        if self.is_drillthrough:
            # Add extra drillthrough param like Store
            base_params.update(self.extra_params)

        for kpi_label, proc_name in sp_map.items():
            # Build EXEC string
            param_list = []
            for key, value in base_params.items():
                if value is None:
                    param_list.append(f"@{key}=NULL")
                elif isinstance(value, str):
                    param_list.append(f"@{key}='{value}'")
                else:
                    param_list.append(f"@{key}={value}")

            exec_string = f"EXEC {proc_name} " + ", ".join(param_list)
            print(f"🧪 Executing: {exec_string}")

            try:
                self.cursor.execute(exec_string)
                result = self.cursor.fetchone()
                if result and len(result) >= 1:
                    db_kpi_data[kpi_label.strip()] = str(result[0]).strip()
                    print(f"✅ {kpi_label}: {str(result[0]).strip()}")
                else:
                    db_kpi_data[kpi_label.strip()] = "No Result"
                    print(f"⚠️ {kpi_label}: No data returned from stored procedure")
            except Exception as e:
                error_msg = str(e)
                if "Could not find stored procedure" in error_msg:
                    print(f"❌ {kpi_label}: Stored procedure '{proc_name}' not found in database")
                    db_kpi_data[kpi_label.strip()] = "SP Not Found"
                else:
                    print(f"❌ {kpi_label}: Database error - {error_msg}")
                    db_kpi_data[kpi_label.strip()] = "DB Error"

        return db_kpi_data

def compare_kpi_data(excel_data, db_data, output_path):
    def normalize_key(key):
        return key.strip().lower().replace(".", "").replace("  ", " ")

    normalized_db_data = {normalize_key(k): v for k, v in db_data.items()}

    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Comparison"
    ws.append(["KPI Name", "Excel Value", "DB Value", "Status", "Notes"])

    matches = 0
    mismatches = 0
    not_found = 0
    sp_errors = 0

    for kpi_name in excel_data:
        excel_value = excel_data[kpi_name]
        normalized_key = normalize_key(kpi_name)
        db_value = normalized_db_data.get(normalized_key, "Not Found")
        
        # Determine status and notes
        if db_value == "Not Found":
            status = "No SP Mapping"
            notes = "No stored procedure mapped for this KPI"
            not_found += 1
        elif db_value == "SP Not Found":
            status = "SP Missing"
            notes = "Stored procedure does not exist in database"
            sp_errors += 1
        elif db_value == "DB Error":
            status = "DB Error"
            notes = "Database error occurred"
            sp_errors += 1
        elif excel_value == db_value:
            status = "Match"
            notes = "Values match perfectly"
            matches += 1
        else:
            status = "Mismatch"
            notes = "Values do not match"
            mismatches += 1
        
        ws.append([kpi_name, excel_value, db_value, status, notes])
        
        # Color-coded console output
        if status == "Match":
            print(f"✅ {kpi_name} -> Excel: {excel_value} | DB: {db_value} => {status}")
        elif status in ["No SP Mapping", "SP Missing"]:
            print(f"⚠️ {kpi_name} -> Excel: {excel_value} | DB: {db_value} => {status}")
        else:
            print(f"❌ {kpi_name} -> Excel: {excel_value} | DB: {db_value} => {status}")

    # Add summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Metric", "Count"])
    ws_summary.append(["Total KPIs", len(excel_data)])
    ws_summary.append(["Matches", matches])
    ws_summary.append(["Mismatches", mismatches])
    ws_summary.append(["No SP Mapping", not_found])
    ws_summary.append(["SP Errors", sp_errors])
    ws_summary.append(["Success Rate", f"{(matches/(len(excel_data)))*100:.1f}%" if excel_data else "0%"])

    wb.save(output_path)
    
    print(f"\n📊 KPI Comparison Summary:")
    print(f"  ✅ Matches: {matches}")
    print(f"  ❌ Mismatches: {mismatches}")
    print(f"  ⚠️ No SP Mapping: {not_found}")
    print(f"  🔧 SP Errors: {sp_errors}")
    print(f"  📈 Success Rate: {(matches/(len(excel_data)))*100:.1f}%" if excel_data else "0%")
    print(f"✅ KPI comparison report saved at: {output_path}")
