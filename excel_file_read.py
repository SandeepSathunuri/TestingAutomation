from openpyxl import load_workbook

def read_kpi_from_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    kpi_dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):  
        kpi_name, kpi_value = row
        kpi_dict[kpi_name.strip()] = str(kpi_value).strip()
    return kpi_dict
